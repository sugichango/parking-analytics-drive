import os
import re
import argparse
from collections import defaultdict
from datetime import datetime, date

import pandas as pd
import jpholiday


# ★ 駐車場ごとの収容台数マスタ ★
PARKING_CAPACITY = {
    "南1駐車場": 918,
    "南１駐車場": 918,
    "南2駐車場": 601,
    "南２駐車場": 601,
    "南3駐車場": 690,
    "南３駐車場": 690,
    "南4駐車場": 638,
    "南４駐車場": 638,
    "南4B駐車場": 0,
    "南４B駐車場": 0,
    "北1駐車場": 622,
    "北１駐車場": 622,
    "北2駐車場": 248,
    "北２駐車場": 248,
    "北3駐車場": 192,
    "北３駐車場": 192,
    "全駐車場": 3809,
}


def parse_filename(filename):
    """
    ファイル名から 日付（YYYYMMDD） と 駐車場名 を取り出す。

    例: '20240401_全駐車場.csv' -> ('20240401', '全駐車場')
    """
    pattern = re.compile(r'(\d{8})_(.+)\.csv$', re.IGNORECASE)
    m = pattern.match(filename)
    if not m:
        return None, None
    date_str, parking_name = m.groups()
    return date_str, parking_name


def sanitize_filename(name):
    """
    Windows のファイル名に使えない文字を '_' に置き換える。
    """
    return re.sub(r'[\\/:*?"<>|]', '_', name)


def collect_csv_by_parking(target_dir):
    """
    指定フォルダ内の CSV を読み込み、
    { 駐車場名: [(date_str, csv_path), ...], ... } という辞書を作る。
    """
    parking_dict = defaultdict(list)

    for fname in os.listdir(target_dir):
        if not fname.lower().endswith(".csv"):
            continue

        date_str, parking_name = parse_filename(fname)
        if date_str is None:
            print(f"[WARN] 想定外のファイル名のためスキップ: {fname}")
            continue

        full_path = os.path.join(target_dir, fname)
        parking_dict[parking_name].append((date_str, full_path))

    # 日付順にソート
    for parking_name in parking_dict:
        parking_dict[parking_name].sort(key=lambda x: x[0])

    return parking_dict


def _convert_mixed_column_to_numeric(col):
    """
    1つの列の中に「数字っぽい文字列」と「日本語など」が混ざっている場合に、
    数字っぽい部分だけ float に変換する。
    ただし列全体が完全に文字列ならそのまま返す。
    """
    if col.dtype != object:
        # すでに数値型などならそのまま返す
        return col

    s = (
        col.astype(str)
        .str.replace("\u3000", " ", regex=False)  # 全角スペース
        .str.strip()
        .str.replace(",", "", regex=False)       # カンマ
    )
    num = pd.to_numeric(s, errors="coerce")

    # 1つも数値に変換できない場合は元の列をそのまま返す
    if num.notna().sum() == 0:
        return col

    # 数値にできた行だけ数値に差し替える
    out = col.copy()
    out.loc[num.notna()] = num.loc[num.notna()]
    return out


def read_csv_flex(csv_path):
    """
    文字コードをいくつか試しながら CSV を読む。
    読めなければ None。
    読めたら「数字っぽい列」は数値型に変換。
    """
    encodings_to_try = ["cp932", "utf-8-sig", "utf-8"]
    last_err = None

    for enc in encodings_to_try:
        try:
            df = pd.read_csv(csv_path, encoding=enc)
            print(f"[INFO] 読み込み成功: {csv_path} (encoding={enc})")

            # 数字っぽい列は数値型へ
            for col_name in df.columns:
                df[col_name] = _convert_mixed_column_to_numeric(df[col_name])

            return df
        except UnicodeDecodeError as e:
            print(f"[WARN] encoding={enc} では読めませんでした: {csv_path} -> {e}")
            last_err = e
        except Exception as e:
            print(f"[WARN] encoding={enc} で別のエラー: {csv_path} -> {e}")
            last_err = e

    print(f"[ERROR] CSV読込に失敗しました: {csv_path}\n  最後のエラー: {last_err}")
    return None


def get_day_type(date_str):
    """
    'YYYYMMDD' から「平日」or「休日」を返す。
    休日 = 土日 or 祝日(jpholiday)
    """
    d = datetime.strptime(date_str, "%Y%m%d").date()
    if d.weekday() >= 5 or jpholiday.is_holiday(d):
        return "休日"
    else:
        return "平日"


def build_monthly_average_dataframe(dfs, year, month,
                                    label_suffix="月平均",
                                    daytype_label="月平均"):
    """
    各日付の DataFrame から「平均」の DataFrame を作る。

    dfs: [(date_str, df), ...]
         各 df は 1日のデータ（日付列/曜日区分列を含む）

    label_suffix:
        「◯年◯月_◯◯」の「◯◯」部分（例: "月平均", "平日平均", "休日平均"）
    daytype_label:
        「曜日区分」列に入れるラベル（例: "月平均", "平日平均", "休日平均"）

    戻り値:
        平均の DataFrame（行構造は1日目の df を踏襲）
        dfs が空の場合は None
    """
    if not dfs:
        return None

    # 1日目の DataFrame をベースにする（行・列の構造を流用）
    base_df = dfs[0][1]
    avg_df = base_df.copy()

    # 列ごとに「全日分の平均」を計算する
    for col in base_df.columns:
        # 日付・曜日区分はあとでラベルとして上書きするのでここでは飛ばす
        if col in ("日付", "曜日区分"):
            continue

        series_list = []
        for _, df in dfs:
            if col not in df.columns:
                continue
            # 文字列が混ざっていても、数値っぽいものだけ取り出す
            s = pd.to_numeric(df[col], errors="coerce")
            series_list.append(s)

        if not series_list:
            continue

        stacked = pd.concat(series_list, axis=1)  # 行×日数 の表
        avg_series = stacked.mean(axis=1, skipna=True)

        # 小数点以下を四捨五入して整数にする
        avg_series = avg_series.round(0)

        # pandas の整数拡張型（欠損も扱える）に変換
        avg_df[col] = avg_series.astype("Int64")

    # ラベル列の整形
    label_for_month = f"{year}年{month:02d}月_{label_suffix}"
    if "日付" in avg_df.columns:
        avg_df["日付"] = label_for_month
    if "曜日区分" in avg_df.columns:
        avg_df["曜日区分"] = daytype_label

    return avg_df


def apply_template_and_chart(target_ws, template_ws, parking_name, year, month):
    """
    target_ws: 平均シート（「月平均」「平日平均」「休日平均」など）
    template_ws: ひな形用シート（最初の日のシート）
    parking_name, year, month: グラフタイトル・D5セル用
    """
    # --- ここから「D列・6〜8行の転記」と「D5書き換え」処理 ---

    max_row_template = template_ws.max_row
    max_col_template = template_ws.max_column

    # ① D列を全てコピー（テンプレートのD列 → target_wsのD列）
    for r in range(1, max_row_template + 1):
        src_cell = template_ws.cell(row=r, column=4)  # D列 = 4
        dst_cell = target_ws.cell(row=r, column=4)
        dst_cell.value = src_cell.value

    # ② 6,7,8行を全てコピー（テンプレートの行 → target_wsの同じ行）
    for r in (6, 7, 8):
        if r > max_row_template:
            continue
        for c in range(1, max_col_template + 1):
            src_cell = template_ws.cell(row=r, column=c)
            dst_cell = target_ws.cell(row=r, column=c)
            dst_cell.value = src_cell.value

    # ③ D5セルを「◯年◯月」に書き換え
    target_ws["D5"].value = f"{year}年{month}月"

    # --- ここからグラフ作成処理 ---
    try:
        from openpyxl.chart import BarChart, LineChart, Reference, Series
        from openpyxl.chart.axis import ChartLines

        # データ行範囲（9〜32行目。データが足りない場合は最終行までに調整）
        start_row = 9
        end_row = min(32, target_ws.max_row)

        if end_row >= start_row:
            # 収容台数（ライン用）
            capacity_value = PARKING_CAPACITY.get(parking_name)

            # 横軸：D列（時間帯など）9〜32行目
            cat_ref = Reference(
                target_ws,
                min_col=4, max_col=4,
                min_row=start_row, max_row=end_row
            )

            # 積み上げ棒グラフ
            bar_chart = BarChart()
            bar_chart.type = "col"
            bar_chart.grouping = "stacked"
            bar_chart.overlap = 100

            # タイトル（Excel デフォルト配置 = 枠外上側）
            bar_chart.title = f"{parking_name}{year}年{month}月平均グラフ"

            # 軸タイトルは非表示（邪魔防止）
            bar_chart.x_axis.title = ""
            bar_chart.y_axis.title = ""

            # 軸ラベル（数値・00-01など）を必ず表示
            bar_chart.x_axis.delete = False
            bar_chart.y_axis.delete = False
            bar_chart.x_axis.tickLblPos = "nextTo"
            bar_chart.y_axis.tickLblPos = "nextTo"

            # 縦軸の数値フォーマット
            bar_chart.y_axis.number_format = "0"

            # 縦軸は 0 以上のみ表示（マイナス不要）
            bar_chart.y_axis.scaling.min = 0

            # ★ 縦軸メモリは「全駐車場だけ 500、他は 100」刻み ★
            if parking_name == "全駐車場":
                bar_chart.y_axis.majorUnit = 500
            else:
                bar_chart.y_axis.majorUnit = 100

            # ★ グリッド線を「線（単色）＋薄いグレー」に強制指定 ★
            try:
                grid = ChartLines()
                ln = grid.graphicalProperties.line

                # 透明度1%の薄いグレー（ARGB形式）
                # A=0x03（約1%）, R,G,B=D0
                ln.solidFill = "03D0D0D0"
                ln.width = 2000  # 細め
                bar_chart.y_axis.majorGridlines = grid
            except Exception as e:
                print(f"[WARN] グリッド線装飾に失敗しました: {e}")

            # 凡例をグラフ下に置き、グラフとは重ねない
            bar_chart.legend.position = "b"
            bar_chart.legend.overlay = False

            # 在庫データ（定期在庫：J列、一般在庫：G列）
            teiki_vals = Reference(
                target_ws,
                min_col=10, max_col=10,
                min_row=start_row, max_row=end_row
            )
            ippan_vals = Reference(
                target_ws,
                min_col=7, max_col=7,
                min_row=start_row, max_row=end_row
            )

            ser_teiki = Series(teiki_vals, title="定期在庫")
            ser_ippan = Series(ippan_vals, title="一般在庫")

            bar_chart.append(ser_teiki)
            bar_chart.append(ser_ippan)

            bar_chart.set_categories(cat_ref)

            # 折れ線グラフ（一般入庫・一般出庫・定期入庫・定期出庫）
            line_chart = LineChart()

            line_series_info = [
                (5, "一般入庫"),
                (6, "一般出庫"),
                (8, "定期入庫"),
                (9, "定期出庫"),
            ]

            for col_idx, series_name in line_series_info:
                vals = Reference(
                    target_ws,
                    min_col=col_idx, max_col=col_idx,
                    min_row=start_row, max_row=end_row
                )
                s = Series(vals, title=series_name)
                line_chart.append(s)

            # ★ 収容台数ライン（赤い点線）を追加 ★
            if capacity_value is not None:
                # T列(20列目)を収容台数用のダミー列として使用
                for r in range(start_row, end_row + 1):
                    target_ws.cell(row=r, column=20, value=capacity_value)

                cap_vals = Reference(
                    target_ws,
                    min_col=20, max_col=20,
                    min_row=start_row, max_row=end_row
                )
                capacity_series = Series(cap_vals, title="収容台数")

                # 線の装飾（赤・点線・少し太め）
                gp_line = capacity_series.graphicalProperties.line
                gp_line.solidFill = "FF0000"      # 赤
                gp_line.dashStyle = "sysDash"     # 点線
                gp_line.width = 30000             # 太さ

                line_chart.append(capacity_series)

            line_chart.set_categories(cat_ref)

            # 棒グラフ + 折れ線グラフ を1つの複合グラフにまとめる
            bar_chart += line_chart

            # 合体後にもう一度カテゴリを設定して「00-01」などを反映
            bar_chart.set_categories(cat_ref)

            # グラフの配置場所（P2付近：表の右側）
            target_ws.add_chart(bar_chart, "P2")

    except Exception as e:
        print(f"[WARN] グラフ作成中にエラーが発生しました: {e}")
    # --- グラフ作成処理ここまで ---


def export_to_excel(parking_dict, output_dir, year, month):
    """
    駐車場ごとの CSV グループから、Excel ファイルを出力する。
    1駐車場 = 1つのExcel
    1日 = 1シート
    各シートの A列=日付, B列=曜日区分（平日/休日）

    ★ 追加仕様 ★
    - 各 Excel に「月平均」シートを1枚追加する
      （全日分の数値列の平均値を計算）
    - さらに「平日平均」「休日平均」シートも追加する
      （それぞれ平日データのみ・休日データのみで平均）
    - 各平均シートに対して、
        ・最初の日のシートから D列全てをコピー
        ・最初の日のシートから 6,7,8行を全てコピー
        ・D5セルを「◯年◯月」に書き換え
      さらに、これまでと同じ形式のグラフを作成する。
    """
    os.makedirs(output_dir, exist_ok=True)

    for parking_name, records in parking_dict.items():
        dfs = []

        for date_str, csv_path in records:
            df = read_csv_flex(csv_path)
            if df is None:
                print(f"[WARN] 読み込めないため、この日付のシートをスキップします: {csv_path}")
                continue

            day_type = get_day_type(date_str)
            dt = datetime.strptime(date_str, "%Y%m%d")

            # ここで必ず先頭2列として「日付」「曜日区分」を追加する
            df = df.copy()
            df.insert(0, "曜日区分", day_type)
            df.insert(0, "日付", dt)

            print(f"[DEBUG] {csv_path} の列名: {list(df.columns)}")

            dfs.append((date_str, df))

        if not dfs:
            print(f"[WARN] 有効なCSVが無かったため、この駐車場のExcelは作成しません: {parking_name}")
            continue

        safe_parking_name = sanitize_filename(parking_name)
        # 既存ファイルと区別するため、末尾に _with_avg を付ける
        excel_filename = f"{year}{month:02d}_{safe_parking_name}_with_avg.xlsx"
        excel_path = os.path.join(output_dir, excel_filename)

        print(f"[INFO] 出力中: {excel_path}  (シート数: {len(dfs)} 日 + 月平均 + 平日平均 + 休日平均)")

        # Excel へ書き出し
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            # 1日1シートを日付順に出力
            for date_str, df in dfs:
                sheet_name = date_str[4:]  # '20240401' -> '0401'
                df.to_excel(writer, sheet_name=sheet_name, index=False)

            # ---- 平均データフレームの作成 ----
            # 全日平均（従来の「月平均」）
            avg_df = build_monthly_average_dataframe(dfs, year, month)

            # 平日だけ
            weekday_dfs = [(d, df) for (d, df) in dfs if get_day_type(d) == "平日"]
            weekday_avg_df = None
            if weekday_dfs:
                weekday_avg_df = build_monthly_average_dataframe(
                    weekday_dfs, year, month,
                    label_suffix="平日平均",
                    daytype_label="平日平均"
                )

            # 休日だけ
            holiday_dfs = [(d, df) for (d, df) in dfs if get_day_type(d) == "休日"]
            holiday_avg_df = None
            if holiday_dfs:
                holiday_avg_df = build_monthly_average_dataframe(
                    holiday_dfs, year, month,
                    label_suffix="休日平均",
                    daytype_label="休日平均"
                )

            # ひな形シート（最初の日）
            wb = writer.book
            first_date_str = dfs[0][0]
            template_sheet_name = first_date_str[4:]  # '20240401' -> '0401'
            template_ws = wb[template_sheet_name]

            # --- 月平均シートを作成（従来の処理） ---
            if avg_df is not None:
                avg_sheet_name = "月平均"
                avg_df.to_excel(writer, sheet_name=avg_sheet_name, index=False)
                avg_ws = wb[avg_sheet_name]
                apply_template_and_chart(avg_ws, template_ws, parking_name, year, month)

            # --- 平日平均シートを作成 ---
            if weekday_avg_df is not None:
                weekday_sheet_name = "平日平均"
                weekday_avg_df.to_excel(writer, sheet_name=weekday_sheet_name, index=False)
                weekday_ws = wb[weekday_sheet_name]
                apply_template_and_chart(weekday_ws, template_ws, parking_name, year, month)

            # --- 休日平均シートを作成 ---
            if holiday_avg_df is not None:
                holiday_sheet_name = "休日平均"
                holiday_avg_df.to_excel(writer, sheet_name=holiday_sheet_name, index=False)
                holiday_ws = wb[holiday_sheet_name]
                apply_template_and_chart(holiday_ws, template_ws, parking_name, year, month)

        print(f"[INFO] 保存完了: {excel_path}")


def main():
    parser = argparse.ArgumentParser(
        description="1ヶ月分の日毎CSVファイルを、駐車場ごとに1つのExcelにまとめ、月平均シート・平日平均シート・休日平均シートも追加するスクリプト"
    )
    parser.add_argument("--year", type=int, required=True)
    parser.add_argument("--month", type=int, required=True)
    parser.add_argument(
        "--csv-base-dir",
        type=str,
        default=r"C:\Users\sugitamasahiko\Documents\parking_system\downloads\csv",
    )
    parser.add_argument(
        "--output-dir",
        type=str,
        default=None,
    )

    args = parser.parse_args()

    target_dir = os.path.join(args.csv_base_dir, f"csv{args.year}{args.month:02d}")
    # 既存と混ざらないよう、デフォルト出力先も _with_avg を付ける
    if args.output_dir is None:
        output_dir = os.path.join(args.csv_base_dir, f"excel{args.year}{args.month:02d}_with_avg")
    else:
        output_dir = args.output_dir

    print(f"[INFO] 対象フォルダ: {target_dir}")
    print(f"[INFO] 出力フォルダ: {output_dir}")

    if not os.path.isdir(target_dir):
        raise FileNotFoundError(f"対象フォルダが存在しません: {target_dir}")

    parking_dict = collect_csv_by_parking(target_dir)

    if not parking_dict:
        print("[WARN] 対象CSVが見つかりませんでした。")
        return

    export_to_excel(parking_dict, output_dir, args.year, args.month)

    print("[INFO] 完了しました。")


if __name__ == "__main__":
    main()
