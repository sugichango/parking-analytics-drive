# -*- coding: utf-8 -*-
"""
make_multi_month_stock_graph_excel改.py

make_monthly_max_stock_csv.py で作成した「***allYYYYMM.csv」を
複数月まとめて結合し、「指定した期間ぜんぶを横軸に並べたグラフ付きExcel」を
駐車場ごとに作成するスクリプト（改良版）。

ポイント:
- 元の make_multi_month_stock_graph_excel.py は一切変更しない
- その中の create_graph_excel_from_csv() を再利用する
  → ファイル名 xxx_graph.xlsx、グラフ別シート、平日／土日祝色分け、
    x軸／y軸ラベルなど、元の単月版と同じ仕様になる
- まず複数月ぶんの CSV を駐車場ごとに縦に結合 → 一時的な結合CSV を保存
- その結合CSV を create_graph_excel_from_csv() に渡してグラフ付きExcelを作る
- その後、「data」シートに
    収容台数・90％・70％ の3列を追加し、
  全グラフに細い折れ線(3本)と説明テキストを追加する
- グラフの x軸・y軸数値を表示 (タイトルは消す)
- 凡例をグラフ枠の外・下に配置

★今回の変更点（重要）:
- graphs_multi_YYYYMM_YYYYMM の「直下」に全てのExcelを出力する（駐車場別サブフォルダを作らない）
- 収容台数ライン追加は「今回新規生成されたxlsx」だけに適用する（出力先共通化による事故防止）
"""

import argparse
import os
import sys
import glob
import re
from collections import defaultdict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import Reference, Series, LineChart

# 既存スクリプトの関数を再利用
from make_monthly_stock_graph_excel import create_graph_excel_from_csv


# ★ 駐車場ごとの収容台数マスタ ★
PARKING_CAPACITY = {
    "南1駐車場": 918,
    "南１駐車場": 918,
    "南2駐車場": 601,
    "南２駐車場": 601,
    "南3駐車場": 690,
    "南３駐車場": 690,
    "南4駐車場": 638,
    "南４駐車場": 638,
    "南4B駐車場": 0,
    "南４B駐車場": 0,
    "北1駐車場": 622,
    "北１駐車場": 622,
    "北2駐車場": 248,
    "北２駐車場": 248,
    "北3駐車場": 192,
    "北３駐車場": 192,
    "全駐車場": 3809,
}


def _to_fullwidth_digits(s: str) -> str:
    """半角数字を全角数字に変換するヘルパー関数。"""
    table = str.maketrans("0123456789", "０１２３４５６７８９")
    return s.translate(table)


def _resolve_capacity_key(pk: str):
    """
    '南1all' → '南1駐車場' などに変換して収容台数を取得する。
    """
    # まずはキーそのもの
    if pk in PARKING_CAPACITY:
        return PARKING_CAPACITY[pk]

    candidates = []

    # 特別扱い: 全駐車場
    if pk == "全駐車場all":
        candidates.append("全駐車場")

    # '◯◯all' → '◯◯駐車場' / 全角数字版
    if pk.endswith("all"):
        base = pk[:-3]
        candidates.append(base + "駐車場")
        candidates.append(_to_fullwidth_digits(base) + "駐車場")
    else:
        candidates.append(pk + "駐車場")
        candidates.append(_to_fullwidth_digits(pk) + "駐車場")

    for name in candidates:
        if name in PARKING_CAPACITY:
            return PARKING_CAPACITY[name]

    return None


def add_capacity_to_excel_charts(xlsx_paths: list[str], parking_key: str) -> None:
    """
    指定されたExcelファイル（複数）に収容台数ライン(100%・90%・70%)を追加する。

    - dataシートに「収容台数」「90％」「70％」列を追加
    - 既存グラフが参照している行範囲と同じ行に値を入れる
    - 3列を参照する LineChart（折れ線）を作成し、
      既存の棒グラフとコンボグラフとして合体させる
    - グラフシートに「収容台数: ○○台」という説明テキストを追加
    - グラフの x軸・y軸の数値ラベルを表示 (タイトルは消去)
    - 凡例をグラフ枠の外・下に配置
    """
    capacity = _resolve_capacity_key(parking_key)
    if capacity is None:
        print(f"[WARN] 収容台数が未定義のため、グラフ修正をスキップします: key={parking_key}")
        return

    cap_90 = round(capacity * 0.9)
    cap_70 = round(capacity * 0.7)

    for xlsx_path in xlsx_paths:
        if not xlsx_path.lower().endswith(".xlsx"):
            continue

        print(f"[INFO] 収容台数ライン追加対象のExcel: {xlsx_path}")

        try:
            wb = load_workbook(xlsx_path)
        except Exception as e:
            print(f"[WARN] Excel読み込みに失敗したためスキップ: {xlsx_path} -> {e}")
            continue

        # dataシートが前提
        if "data" not in wb.sheetnames:
            print(f"[WARN] 'data' シートが見つからないためスキップ: {xlsx_path}")
            continue

        data_ws = wb["data"]

        # グラフを持つシート（通常は graph シート）を探す
        graph_sheets = []
        for ws in wb.worksheets:
            charts = getattr(ws, "_charts", None)
            if charts is None:
                charts = getattr(ws, "charts", [])
            if charts:
                graph_sheets.append(ws)

        if not graph_sheets:
            print(f"[WARN] グラフを持つシートが無いためスキップ: {xlsx_path}")
            continue

        # 代表として、最初のグラフシートの最初の系列から行範囲を取得
        charts0 = getattr(graph_sheets[0], "_charts", None)
        if charts0 is None:
            charts0 = getattr(graph_sheets[0], "charts", [])
        if not charts0:
            print(f"[WARN] グラフが無いためスキップ: {xlsx_path}")
            continue

        first_chart = charts0[0]
        if not first_chart.series:
            print(f"[WARN] グラフに系列が無いためスキップ: {xlsx_path}")
            continue

        first_series = first_chart.series[0]

        # 例: "data!$P$9:$P$245" から行番号 9〜245 を抜き出す
        try:
            ref_str = first_series.val.numRef.f  # "data!$P$9:$P$245" など
        except Exception as e:
            print(f"[WARN] 参照範囲の取得に失敗したためスキップ: {xlsx_path} -> {e}")
            continue

        m = re.search(r"\$(\w+)\$(\d+):\$(\w+)\$(\d+)", ref_str)
        if not m:
            print(f"[WARN] 行範囲を解釈できなかったためスキップ: {xlsx_path} (ref={ref_str})")
            continue

        start_row = int(m.group(2))
        end_row = int(m.group(4))
        print(f"[INFO] データ行範囲: {start_row}〜{end_row}")

        # --- dataシートに「収容台数」「90％」「70％」の3列を追加 ---
        cap_col = data_ws.max_column + 1
        cap90_col = cap_col + 1
        cap70_col = cap_col + 2

        data_ws.cell(row=1, column=cap_col, value="収容台数")
        data_ws.cell(row=1, column=cap90_col, value="90％")
        data_ws.cell(row=1, column=cap70_col, value="70％")

        for r in range(start_row, end_row + 1):
            data_ws.cell(row=r, column=cap_col, value=capacity)
            data_ws.cell(row=r, column=cap90_col, value=cap_90)
            data_ws.cell(row=r, column=cap70_col, value=cap_70)

        cap_ref = Reference(
            data_ws, min_col=cap_col, max_col=cap_col, min_row=start_row, max_row=end_row
        )
        cap90_ref = Reference(
            data_ws, min_col=cap90_col, max_col=cap90_col, min_row=start_row, max_row=end_row
        )
        cap70_ref = Reference(
            data_ws, min_col=cap70_col, max_col=cap70_col, min_row=start_row, max_row=end_row
        )

        modified = False

        # --- 各グラフに「収容台数・90％・70％」折れ線をコンボで追加 ---
        for ws in graph_sheets:
            charts = getattr(ws, "_charts", None)
            if charts is None:
                charts = getattr(ws, "charts", [])
            if not charts:
                continue

            # 収容台数の説明テキスト（お好みで位置は調整可）
            if ws["B2"].value in (None, ""):
                ws["B2"].value = f"収容台数: {capacity}台"

            for chart in charts:
                # 軸タイトルを消し、数値ラベルを表示
                try:
                    chart.x_axis.title = ""
                    chart.y_axis.title = ""
                    chart.x_axis.tickLblPos = "nextTo"
                    chart.y_axis.tickLblPos = "nextTo"
                    chart.x_axis.delete = False
                    chart.y_axis.delete = False
                except Exception:
                    pass

                # 凡例を枠の外・下に
                try:
                    chart.legend.position = "b"   # bottom
                    chart.legend.overlay = False  # グラフと重ねない
                except Exception:
                    pass

                # グラフサイズを大きくする
                try:
                    chart.width = 40
                    chart.height = 25
                except Exception:
                    pass

                # 2回目以降の重複追加防止：すでに「収容台数」系列があればスキップ
                has_capacity = False
                for s in chart.series:
                    try:
                        t = s.title
                        if hasattr(t, "v"):
                            title_text = str(t.v)
                        else:
                            title_text = str(t)
                    except Exception:
                        title_text = ""
                    if "収容台数" in title_text:
                        has_capacity = True
                        break

                if has_capacity:
                    continue

                # 折れ線グラフ(LineChart)を作成し、3本の系列を追加
                line_chart = LineChart()

                ser_cap = Series(cap_ref, title="収容台数")
                ser_90 = Series(cap90_ref, title="90％")
                ser_70 = Series(cap70_ref, title="70％")

                # 線の装飾（細め・点線・色分け・マーカー無し）
                for ser, color in [
                    (ser_cap, "FF0000"),   # 赤
                    (ser_90, "FFA500"),    # オレンジ
                    (ser_70, "008000"),    # 緑
                ]:
                    line = ser.graphicalProperties.line
                    line.solidFill = color
                    line.dashStyle = "sysDash"
                    line.width = 20000
                    try:
                        ser.marker.symbol = "none"
                    except Exception:
                        pass

                line_chart.append(ser_cap)
                line_chart.append(ser_90)
                line_chart.append(ser_70)

                # 既存の棒グラフとコンボにする
                chart += line_chart
                modified = True

        if modified:
            try:
                wb.save(xlsx_path)
                print(f"[INFO] 収容台数ライン(100/90/70%)を追加して保存しました: {xlsx_path}")
            except Exception as e:
                print(f"[WARN] 収容台数ライン追加後の保存に失敗しました: {xlsx_path} -> {e}")
        else:
            print(f"[INFO] 収容台数系列の追加は不要でした: {xlsx_path}")


def iterate_year_month(start_year: int, start_month: int, end_year: int, end_month: int):
    """
    start_year, start_month 〜 end_year, end_month の範囲の (year, month) を
    1ヶ月ずつ増やしながら順番に返すジェネレータ。
    例）2025-04 〜 2025-06
        -> (2025, 4), (2025, 5), (2025, 6)
    """
    y, m = start_year, start_month
    while (y < end_year) or (y == end_year and m <= end_month):
        yield y, m
        m += 1
        if m > 12:
            m = 1
            y += 1


def main():
    parser = argparse.ArgumentParser(
        description="***allYYYYMM.csv を複数月まとめて結合し、期間全体のグラフ付きExcelを作成するスクリプト（改）"
    )
    parser.add_argument("--start-year", type=int, required=True, help="開始年（例: 2025）")
    parser.add_argument("--start-month", type=int, required=True, help="開始月（1〜12）")
    parser.add_argument("--end-year", type=int, required=True, help="終了年（例: 2025）")
    parser.add_argument("--end-month", type=int, required=True, help="終了月（1〜12）")
    parser.add_argument(
        "--base-dir",
        type=str,
        default=r"downloads\csv",
        help="csvYYYYMM フォルダが入っているベースディレクトリ（既定: downloads\\csv）",
    )
    parser.add_argument("--date-col", type=str, default="日付", help="日付列の列名（既定: '日付'）")
    parser.add_argument("--general-col", type=str, default="一般在庫", help="一般在庫列の列名（既定: '一般在庫'）")
    parser.add_argument("--teiki-col", type=str, default="定期在庫", help="定期在庫列の列名（既定: '定期在庫'）")
    parser.add_argument(
        "--holiday-col", type=str, default="土日祝区分", help="土日祝区分の列名（既定: '土日祝区分'）"
    )
    parser.add_argument(
        "--output-subdir",
        type=str,
        default="graphs_multi",
        help="結合期間のグラフ付きExcelを出力するサブフォルダ名（既定: 'graphs_multi'）",
    )
    parser.add_argument(
        "--temp-subdir",
        type=str,
        default="csv_multi",
        help="結合後CSVを保存するサブフォルダ名（既定: 'csv_multi'）",
    )

    args = parser.parse_args()

    # 開始>終了のときはエラー
    if (args.start_year > args.end_year) or (
        args.start_year == args.end_year and args.start_month > args.end_month
    ):
        print("[ERROR] 開始年月が終了年月より後になっています。指定を見直してください。")
        sys.exit(1)

    period_str = f"{args.start_year}{args.start_month:02d}_{args.end_year}{args.end_month:02d}"
    print("[INFO] === 複数月結合グラフ作成を開始します（改） ===")
    print(f"[INFO] 対象期間: {args.start_year}-{args.start_month:02d} 〜 {args.end_year}-{args.end_month:02d}")
    print(f"[INFO] ベースディレクトリ: {args.base_dir}")

    # 駐車場ごとのファイル一覧を溜める {key: [csv_path, ...]}
    # key は「南1all」「全駐車場all」など、「allYYYYMM.csv」から YYYYMM を除いた部分
    files_by_key: dict[str, list[str]] = defaultdict(list)

    # まずは、指定期間のすべての ***allYYYYMM.csv を探して、駐車場ごとにまとめる
    for year, month in iterate_year_month(args.start_year, args.start_month, args.end_year, args.end_month):
        ym = f"{year}{month:02d}"
        monthly_dir = os.path.join(args.base_dir, f"csv{ym}")
        if not os.path.isdir(monthly_dir):
            print(f"[WARN] 月次フォルダが存在しません（スキップ）: {monthly_dir}")
            continue

        pattern = os.path.join(monthly_dir, f"*all{ym}.csv")
        csv_files = glob.glob(pattern)

        if not csv_files:
            print(f"[WARN] 月次まとめCSVが見つかりませんでした（スキップ）: {pattern}")
            continue

        print(f"[INFO] {ym} の対象CSVファイル数: {len(csv_files)}")

        for csv_path in csv_files:
            base_name = os.path.basename(csv_path)
            # 例: "南1all202510.csv" → key="南1all"
            m = re.match(r"(.+all)\d{6}\.csv$", base_name)
            if not m:
                print(f"[WARN] 予期しないファイル名のためスキップ: {base_name}")
                continue
            key = m.group(1)
            files_by_key[key].append(csv_path)

    if not files_by_key:
        print("[ERROR] 対象期間内に ***allYYYYMM.csv が1件も見つかりませんでした。")
        sys.exit(1)

    # 結合後CSVを保存するフォルダ
    temp_root_dir = os.path.join(args.base_dir, args.temp_subdir + "_" + period_str)
    os.makedirs(temp_root_dir, exist_ok=True)

    # グラフ付きExcelの出力フォルダ（★この直下に全部出す★）
    output_root_dir = os.path.join(args.base_dir, args.output_subdir + "_" + period_str)
    os.makedirs(output_root_dir, exist_ok=True)

    # 駐車場ごとに結合 → 一時CSV作成 → 既存関数でグラフ付きExcel作成 → 収容台数ライン追加
    for key, paths in sorted(files_by_key.items()):
        print("===================================================")
        print(f"[INFO] 駐車場グループ: {key}")
        print(f"[INFO] 対象ファイル数: {len(paths)}")

        # 月順に並ぶよう、パス名でソート（ファイル名に YYYYMM が入っているので文字列ソートでOK）
        paths_sorted = sorted(paths)

        dfs = []
        for p in paths_sorted:
            print(f"[INFO] 読み込み: {p}")
            df = pd.read_csv(p, encoding="cp932")
            dfs.append(df)

        if not dfs:
            print(f"[WARN] 結合対象データがありません（スキップ）: {key}")
            continue

        df_all = pd.concat(dfs, ignore_index=True)

        # 日付でソートしておく（文字列でも良いが、念のため datetime にしてからソート）
        if args.date_col not in df_all.columns:
            print(f"[ERROR] 日付列 '{args.date_col}' が見つかりません（スキップ）: {key}")
            continue

        try:
            dt = pd.to_datetime(df_all[args.date_col])
        except Exception as e:
            print(f"[WARN] 日付列の変換に失敗しましたが、そのまま文字列としてソートします: {e}")
            df_all = df_all.sort_values(args.date_col)
        else:
            df_all[args.date_col] = dt
            df_all = df_all.sort_values(args.date_col)
            # CSVに書き出すときは YYYY-MM-DD の文字列に戻しておく
            df_all[args.date_col] = df_all[args.date_col].dt.strftime("%Y-%m-%d")

        # ファイル名用に Windows で使えない文字を避ける（ファイル名のみサニタイズ）
        safe_key = re.sub(r"[\\/:*?\"<>|]", "_", key)
        combined_name = f"{safe_key}{period_str}.csv"
        combined_path = os.path.join(temp_root_dir, combined_name)

        print(f"[INFO] 結合CSV出力: {combined_path}")
        df_all.to_csv(combined_path, index=False, encoding="cp932")

        # ★出力はすべて output_root_dir 直下に統一する（サブフォルダは作らない）★
        output_dir = output_root_dir

        # create_graph_excel_from_csv が作った xlsx だけを特定するため、前後差分を取る
        before = set(glob.glob(os.path.join(output_dir, "*.xlsx")))

        print(f"[INFO] グラフ付きExcel作成（出力先: {output_dir}）")
        create_graph_excel_from_csv(
            csv_path=combined_path,
            date_col=args.date_col,
            general_col=args.general_col,
            teiki_col=args.teiki_col,
            holiday_col=args.holiday_col,
            output_dir=output_dir,
        )

        after = set(glob.glob(os.path.join(output_dir, "*.xlsx")))
        new_files = sorted(list(after - before))

        if not new_files:
            print(f"[WARN] 新規に作成されたxlsxが見つかりませんでした: key={key}")
        else:
            # ★この駐車場で新規作成されたExcelだけに収容台数ラインを追加★
            add_capacity_to_excel_charts(xlsx_paths=new_files, parking_key=key)

    print("[INFO] === すべての結合グラフ付きExcel作成が完了しました（改） ===")


if __name__ == "__main__":
    main()
