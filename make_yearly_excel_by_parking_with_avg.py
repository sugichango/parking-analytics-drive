import os
import shutil
import argparse
import calendar
from datetime import date
import openpyxl
import jpholiday

def get_month_days(year, month):
    """
    指定された年月の、全日数、平日日数、休日日数を計算する。
    休日は土日およびjpholidayの祝日。
    """
    total_days = calendar.monthrange(year, month)[1]
    weekdays = 0
    holidays = 0
    
    for day in range(1, total_days + 1):
        dt = date(year, month, day)
        if dt.weekday() >= 5 or jpholiday.is_holiday(dt):
            holidays += 1
        else:
            weekdays += 1
            
    return total_days, weekdays, holidays

def extract_parking_names(csv_base_dir, months):
    """
    対象月の excelYYYYMM_with_avg フォルダから、存在するすべての駐車場名を収集する。
    """
    parking_names = set()
    for y, m in months:
        target_dir = os.path.join(csv_base_dir, f"excel{y}{m:02d}_with_avg")
        if os.path.isdir(target_dir):
            for fname in os.listdir(target_dir):
                if fname.endswith("_with_avg.xlsx"):
                    # {yyyy}{mm}_{parking}_with_avg.xlsx から駐車場名を抽出
                    name_part = fname.replace("_with_avg.xlsx", "")
                    if "_" in name_part:
                        p_name = name_part.split("_", 1)[1]
                        parking_names.add(p_name)
    return parking_names

def main():
    parser = argparse.ArgumentParser(
        description="複数月の月平均Excelを加重串刺し平均し、年度版Excelを生成するスクリプト"
    )
    parser.add_argument("--start-year", type=int, required=True)
    parser.add_argument("--start-month", type=int, required=True)
    parser.add_argument("--end-year", type=int, required=True)
    parser.add_argument("--end-month", type=int, required=True)
    parser.add_argument(
        "--csv-base-dir",
        type=str,
        default=r"C:\Users\sugitamasahiko\Documents\parking_system\downloads\csv",
        help="excelYYYYMM_with_avg が入っているベースフォルダ",
    )
    parser.add_argument(
        "--output-dir",
        type=str,
        default=None,
        help="省略した場合は excel{start_year}_with_avg フォルダになります",
    )

    args = parser.parse_args()

    # 対象月のリストを作成
    months = []
    y, m = args.start_year, args.start_month
    while (y < args.end_year) or (y == args.end_year and m <= args.end_month):
        months.append((y, m))
        m += 1
        if m > 12:
            y += 1
            m = 1

    if not months:
        print("[ERROR] 開始年月が終了年月より後になっています。")
        return

    # 全駐車場のリストを取得
    parking_names = extract_parking_names(args.csv_base_dir, months)
    
    if not parking_names:
        print("[WARN] 対象となるExcelファイルが見つかりませんでした。")
        return

    # 出力フォルダの設定
    if args.output_dir is None:
        out_dir = os.path.join(args.csv_base_dir, f"excel{args.start_year}_with_avg")
    else:
        out_dir = args.output_dir
        
    os.makedirs(out_dir, exist_ok=True)
    print(f"[INFO] 出力フォルダ: {out_dir}")

    sheet_types = ["月平均", "平日平均", "休日平均"]

    # 駐車場ごとに処理
    for p_name in sorted(list(parking_names)):
        print(f"[INFO] {p_name} の処理を開始します...")
        
        # 累積データを格納する辞書
        accum_data = {
            st: {"days": 0, "sum": {}} for st in sheet_types
        }
        
        template_file_path = None

        # 各月のファイルを読み込み、数値を日数倍して加算する
        for y, m in months:
            total_days, weekdays, holidays = get_month_days(y, m)
            weights = {
                "月平均": total_days,
                "平日平均": weekdays,
                "休日平均": holidays
            }
            
            src_file = os.path.join(args.csv_base_dir, f"excel{y}{m:02d}_with_avg", f"{y}{m:02d}_{p_name}_with_avg.xlsx")
            if not os.path.exists(src_file):
                continue
                
            if template_file_path is None:
                template_file_path = src_file
                
            # data_only=True で数式ではなく計算済みの値を読み込む
            wb = openpyxl.load_workbook(src_file, data_only=True)
            
            for st in sheet_types:
                if st in wb.sheetnames:
                    ws = wb[st]
                    wt = weights[st]
                    
                    if wt > 0:
                        accum_data[st]["days"] += wt
                        
                        # 9行目～35行目、E列(5)～T列(20)の数値を加算
                        for r in range(9, 36):
                            for c in range(5, 22):
                                try:
                                    val = ws.cell(row=r, column=c).value
                                    if isinstance(val, (int, float)):
                                        key = (r, c)
                                        if key not in accum_data[st]["sum"]:
                                            accum_data[st]["sum"][key] = 0
                                        accum_data[st]["sum"][key] += val * wt
                                except:
                                    pass
            wb.close()
            
        if template_file_path is None:
            print(f"[WARN] {p_name} の対象ファイルが1つも存在しませんでした。スキップします。")
            continue
            
        # --- 新しいファイルを作成 ---
        out_file = os.path.join(out_dir, f"{args.start_year}_{p_name}_with_avg.xlsx")
        
        # テンプレートに用いるエクセルをそのままコピーする
        shutil.copy(template_file_path, out_file)
        
        # コピーしたファイルを開いて編集する
        wb_out = openpyxl.load_workbook(out_file)
        
        # 不要なシート（日別シートなど）の削除
        for s_name in wb_out.sheetnames:
            if s_name not in sheet_types:
                wb_out.remove(wb_out[s_name])
                
        # 計算した加重平均値の上書きと、タイトルの変更
        label_period = f"{args.start_year}年度"
        
        for st in sheet_types:
            if st in wb_out.sheetnames:
                ws = wb_out[st]
                total_days_acc = accum_data[st]["days"]
                
                # 数値の上書き
                if total_days_acc > 0:
                    for (r, c), sum_val in accum_data[st]["sum"].items():
                        # 平均を算出し、四捨五入して整数にする
                        avg_val = round(sum_val / total_days_acc)
                        ws.cell(row=r, column=c).value = avg_val
                
                # D5セルを「2023年度」などに変更
                ws["D5"].value = label_period
                
                # グラフタイトルの変更
                chart_title_str = f"{p_name}{label_period}{st}グラフ"
                for chart in getattr(ws, "_charts", []):
                    # openpyxlでグラフタイトルを設定
                    chart.title = chart_title_str
                    
        # 保存
        wb_out.save(out_file)
        wb_out.close()
        
        print(f"[INFO] 作成完了: {out_file} (重み合計: 全日={accum_data['月平均']['days']}日, 平日={accum_data['平日平均']['days']}日, 休日={accum_data['休日平均']['days']}日)")

   \\10.0.1.113\企画・建設\8.管制システム\R8年度(2026年度)\202604_理事長説明資料

if __name__ == "__main__":
    main()
